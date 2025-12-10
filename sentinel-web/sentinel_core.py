
import os
import re
import math
import random
import warnings
import calendar
import unicodedata
import json
from typing import Dict, Tuple, Any
import numpy as np
import pandas as pd
from ortools.sat.python import cp_model
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# ------------------------
# 0) 사용자 기본 설정
# ------------------------
YEAR, MONTH = 2025, 12 #수정가능
DAYS = calendar.monthrange(YEAR, MONTH)[1]
DAY_SIZE, NIGHT_SIZE = 6, 6




FILE_SQUADS = "분대편성표.xlsx"
FILE_VAC    = "12월 휴가.xlsx"
FILE_GUNS   = "총기.xlsx"
FILE_RANK   = "짬표.xlsx"
OUT_PATH    = f"{YEAR}년_{MONTH:02d}월_공정표_경작서.xlsx"

# ------------------------
# 1) 유틸 / 로더(절대 손대지 말것)
# ------------------------


def prev_year_month(y: int, m: int) -> tuple[int, int]:
    if m == 1:
        return (y - 1, 12)
    return (y, m - 1)

def load_rolling_context(path: str):
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _parse_shift_label(tag: str) -> tuple[str | None, int | None]:
    """주간/야간/예비/휴가를 D/N/R/V로, 조(A/B/C)는 0/1/2로 인코딩."""
    t = str(tag or "").strip()
    if t.startswith("주간"):
        # 라벨 찾기
        for L, idx in zip(["C","B","A"], [0,1,2]):
            if f"{L}조" in t:
                return "D", idx
        return "D", None
    if t.startswith("야간"):
        for L, idx in zip(["C","B","A"], [0,1,2]):
            if f"{L}조" in t:
                return "N", idx
        return "N", None
    if t == "예비":  return "R", None
    if t == "휴가":  return "V", None
    return None, None



def _infer_last_shift_from_trw(trw: list) -> int|None:
    # d 내림차순: D/N이 처음 보이면 그걸 last_shift로
    for rec in sorted(trw or [], key=lambda r: r["d"], reverse=True):
        s = rec.get("shift")
        if s == "D": return 0
        if s == "N": return 1
    return None

def _infer_last_label_from_trw(trw: list, target_shift_char: str) -> int|None:
    # d 내림차순: target 교대(D/N)에서 label이 처음 보이면 채택
    for rec in sorted(trw or [], key=lambda r: r["d"], reverse=True):
        if rec.get("shift") == target_shift_char and rec.get("label") is not None:
            return int(rec["label"])
    return None



def save_rolling_context_from_schedule(path: str, schedule: Dict[int, Dict[str,str]],
                                       members: list, year: int, month: int, days: int):
    """
    solver, x/y/r 없이도 '월간 시트 텍스트'만으로
    다음달에 필요한 최소 정보(마지막날 교대/라벨)만 저장
    """
    ctx = {"year": year, "month": month, "members": members, "per_member": {}}
    last_day = schedule.get(days, {})
    label2idx = {'C': 0, 'B': 1, 'A': 2}

    start_d=max(1, days-6)
    trailing_days = list(range(start_d, days + 1))

    fairness_map = compute_excel_style_fairness(schedule, members, year, month, days)


    for name in members:
        tag = str(last_day.get(name, "")).strip()
        # 마지막 교대(주/야) 판별
        if tag.startswith("주간"):
            last_shift = 0
        elif tag.startswith("야간"):
            last_shift = 1
        else:
            last_shift = None

        last_label = None
        for L in ['C', 'B', 'A']:
            if f"{L}조" in tag:
                last_label = label2idx[L]
                break

        tag = str(last_day.get(name, "")).strip()
        trw = []
        for d in trailing_days:
          t = schedule[d].get(name, "")
          s, ell = _parse_shift_label(t)
          trw.append({
                "d": int(d),
                "shift": s,
                "label": (int(ell) if ell is not None else None),
                "is_weekend": int(calendar.weekday(year, month, d) >= 5),
            })

        if tag.startswith("주간"):
           last_shift = 0
        elif tag.startswith("야간"):
            last_shift = 1
        else:
            last_shift = None


        label2idx = {'C': 0, 'B': 1, 'A': 2}
        last_label = None
        for L in ['C','B','A']:
            if f"{L}조" in tag:
                last_label = label2idx[L]
                break

    # 4) 보정: last_shift가 None이면 trw로 추론
        if last_shift not in (0, 1):
            last_shift = _infer_last_shift_from_trw(trw)

        # 5) 보정: last_label이 None이면 trw 기반으로 해당 교대의 최근 라벨 추론
        if last_shift in (0, 1) and last_label not in (0, 1, 2):
            want = "D" if last_shift == 0 else "N"
            last_label = _infer_last_label_from_trw(trw, want)


        ctx["per_member"][name] = {"last_shift": last_shift, "last_label": last_label, "trailing_week": trw, "prev_fairness": fairness_map.get(name, None) }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(ctx, f, ensure_ascii=False, indent=2)



def normalize_name(s):
    return re.sub(r"\s+"," ",unicodedata.normalize("NFC",str(s or "")).strip())


def canonical_role(v: str) -> str:
    t = str(v or "").strip().lower()
    if t in ("사수", "senior", "leader"): return "leader"
    if t in ("부사수", "assistant", "asso", "assist"): return "assistant"
    return ""

def safe_norm_map(d: dict) -> dict:
    # 이름 키는 normalize, 값은 역할 표준화
    return {normalize_name(k): canonical_role(v) for k, v in (d or {}).items()}


def _norm_cols(df):
    df.columns=[str(c).strip() for c in df.columns]
    return df

def is_weekend(y,m,d):
    return calendar.weekday(y,m,d) >= 5

def load_members(path):
    df=pd.read_excel(path); df=_norm_cols(df)
    if "이름" not in df.columns:
        for c in ["성명","Name","name"]:
            if c in df.columns: df.rename(columns={c:"이름"}, inplace=True)
    names=(df["이름"].astype(str).map(normalize_name).replace({"nan":""}).tolist())
    return list(dict.fromkeys([n for n in names if n]))

def parse_vacation_sheet(path,y,m,days):
    vac=set(); df=pd.read_excel(path); df=_norm_cols(df)
    if "이름" not in df.columns:
        for c in ["성명","Name","name"]:
            if c in df.columns: df.rename(columns={c:"이름"}, inplace=True)
    df["이름"]=df["이름"].map(normalize_name)
    day_cols=[str(d) for d in range(1,days+1)]
    if all(c in df.columns for c in day_cols):
        for _,r in df.iterrows():
            for d in range(1,days+1):
                if str(r[str(d)]).strip().lower() not in ("","nan","none"):
                    vac.add((r["이름"],d))
        return vac
    if "시작일" in df.columns and "종료일" in df.columns:
        df["시작일"]=pd.to_datetime(df["시작일"],errors="coerce")
        df["종료일"]=pd.to_datetime(df["종료일"],errors="coerce")
        for _,r in df.iterrows():
            if pd.notna(r["시작일"]) and pd.notna(r["종료일"]):
                rs=max(r["시작일"].normalize(),pd.Timestamp(y,m,1))
                re=min(r["종료일"].normalize(),pd.Timestamp(y,m,days))
                for d in range(rs.day,re.day+1): vac.add((r["이름"],d))
        return vac
    if "날짜" in df.columns:
        df["날짜"]=pd.to_datetime(df["날짜"],errors="coerce")
        for _,r in df.iterrows():
            if pd.notna(r["날짜"]) and r["날짜"].year==y and r["날짜"].month==m:
                vac.add((r["이름"],r["날짜"].day))
        return vac
    return vac

def load_k15_set(path):
    try:
        df=pd.read_excel(path,header=None); gun_owners={}
        for r in range(df.shape[0]):
            gun=str(df.iloc[r,0]).strip().lower()
            owners=df.iloc[r,1:].dropna().astype(str).map(normalize_name).tolist()
            gun_owners[gun]=owners
        return set(gun_owners.get("k15",[]))|set(gun_owners.get("K15",[]))
    except Exception:
        return set()

def load_rank_priority(path):
    try:
        df = pd.read_excel(path)
    except Exception:
        raise
    df = _norm_cols(df)
    if "이름" not in df.columns:
        for c in ["성명","Name","name"]:
            if c in df.columns:
                df.rename(columns={c:"이름"}, inplace=True)
                break
    if "이름" not in df.columns:
        raise ValueError("짬표 파일에 '이름' 컬럼이 필요합니다.")
    ordered = [normalize_name(v) for v in df["이름"].dropna().astype(str).tolist()]
    return {name: idx for idx, name in enumerate(ordered)}

def load_rankbook(path):
    """
    '짬표.xlsx'에 열이 '사수', '부사수'로 나뉜 포맷을 읽는다.
    반환:
      - role_by_name: {이름: "사수"|"부사수"}
      - leader_order: {이름: 사수열에서의 순서(0이 위)}
      - assistant_order: {이름: 부사수열에서의 순서}
      - display_priority: 표시/정렬용 가중치 (작을수록 왼쪽)
    """
    df = pd.read_excel(path)
    df = _norm_cols(df)

    # 컬럼명 방어
    candidates_lead  = [c for c in df.columns if "사수" in c]
    candidates_assi  = [c for c in df.columns if "부사수" in c]
    if not candidates_lead or not candidates_assi:
        raise ValueError("짬표 파일에 '사수'와 '부사수' 열이 필요합니다.")

    col_lead = candidates_lead[0]
    col_assi = candidates_assi[0]

    leaders    = [normalize_name(v) for v in df[col_lead].dropna().astype(str).tolist()]
    assistants = [normalize_name(v) for v in df[col_assi].dropna().astype(str).tolist()]

    role_by_name   = {}
    leader_order   = {}
    assistant_order= {}

    for i, nm in enumerate(leaders):
        if nm:
            role_by_name[nm] = "사수"
            leader_order[nm] = i
    for i, nm in enumerate(assistants):
        if nm:
            # 사수/부사수 중복이면 '사수' 우선
            role_by_name.setdefault(nm, "부사수")
            if role_by_name.get(nm) == "부사수":
                assistant_order[nm] = i

    # 표시용 우선순위: 사수 먼저(작은 값), 부사수는 큰 오프셋을 더해 뒤로
    display_priority = {}
    for nm in leaders:
        display_priority[nm] = leader_order.get(nm, 10**6)
    for nm in assistants:
        display_priority.setdefault(nm, 10**6 + assistant_order.get(nm, 10**6))


    return {
        "role_by_name": role_by_name,
        "leader_order": leader_order,
        "assistant_order": assistant_order,
        "display_priority": display_priority,
    }



def compute_excel_style_fairness(schedule: Dict[int, Dict[str,str]], members: list,
                                 year:int, month:int, days:int,
                                 w_ratio: float = 0.1,  # 3순위: 주/야 비율(주간 비중 높을수록 +)
                                 w_res: float   = 0.4, # 2순위: 총 예비일수(많을수록 +)
                                 w_low: float   = 0.5  # 1순위: 총 근무일수(적을수록 +)
                                 ) -> Dict[str, float]:
    """
    공정성 점수(0~100, 높을수록 좋음).
      - ratio_score  = day / (day + night)      ; (근무가 0이면 0.5로 중립)
      - reserve_score= reserve / max_reserve    ; (월내 최대 예비수로 정규화)
      - lowload_score= 1 - total / max_total    ; (월내 최대 근무합 대비 적을수록 +
    최종 = 100 * ( w_ratio*ratio_score + w_res*reserve_score + w_low*lowload_score )
    """
    def parse_shift(tag: str):
        t = str(tag or "").strip()
        if t.startswith("주간"): return "D"
        if t.startswith("야간"): return "N"
        if t == "예비":          return "R"
        if t == "휴가":          return "V"
        return ""

    day_cnt, night_cnt, res_cnt, total_cnt = {}, {}, {}, {}
    for m in members:
        seq = [parse_shift(schedule[d].get(m, "")) for d in range(1, days+1)]
        d = sum(s == "D" for s in seq)
        n = sum(s == "N" for s in seq)
        r = sum(s == "R" for s in seq)
        t = d + n
        day_cnt[m], night_cnt[m], res_cnt[m], total_cnt[m] = d, n, r, t

    max_res = max(res_cnt.values()) if res_cnt else 0
    max_tot = max(total_cnt.values()) if total_cnt else 0

    fair = {}
    for m in members:
        d, n, r, t = day_cnt[m], night_cnt[m], res_cnt[m], total_cnt[m]
        ratio_score  = (d / (d + n)) if (d + n) > 0 else 0.5
        reserve_score= (r / max_res) if max_res > 0 else 0.0
        lowload_score= (1.0 - (t / max_tot)) if max_tot > 0 else 1.0

        score = 100.0 * (w_ratio*ratio_score + w_res*reserve_score + w_low*lowload_score)
        fair[m] = float(round(max(0.0, min(100.0, score)), 0))
    return fair




def load_schedule_from_excel(path: str, members: list, days: int) -> Dict[int, Dict[str, str]]:
    """
    export_schedule_to_excel(...)로 만든 파일의 '월간' 시트에서
    {day: {이름: "주간 C조"/"야간 A조"/"예비"/"휴가"}} 형태로 역파싱.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"스케줄 파일이 없습니다: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb["월간"] if "월간" in wb.sheetnames else wb.active

    # 1행: "이름", "1","2",...,"days"
    # A열: 이름
    sched = {d: {} for d in range(1, days + 1)}

    # 이름을 정규화해서 키를 맞춘다
    member_set = {normalize_name(m) for m in members}

    for r in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=r, column=1).value
        name = normalize_name(raw_name)
        if not name or name not in member_set:
            continue
        for d in range(1, days + 1):
            val = ws.cell(row=r, column=1 + d).value
            tag = str(val).strip() if val is not None else ""
            # export로 만든 파일이면 공란이 거의 없지만, 혹시 모르면 공란은 "예비"로 치환
            if tag == "":
                tag = "예비"
            sched[d][name] = tag
    return sched



# === 기존 스케줄에서 휴가 set 뽑기 ===
def vac_from_schedule(schedule: Dict[int, Dict[str,str]], members: list, days: int):
    vac = set()
    for d in range(1, days+1):
        for m in members:
            if str(schedule[d].get(m, "")).strip() == "휴가":
                vac.add((normalize_name(m), d))
    return vac

# === 락-정원/휴가 충돌 사전 점검 ===
def debug_lock_conflicts(hints, schedule, members, days, day_size, night_size, vac_effect):
    name2mi = {normalize_name(n): i for i,n in enumerate(members)}
    # 1) 휴가/교육인데 x=1로 락된 경우
    bad = []
    for (mi,d,s), val in (hints or {}).get("locks", {}).get("x", {}).items():
        if val==1:
            nm = members[mi]
            if (normalize_name(nm), d) in vac_effect:
                bad.append((d, "휴가/교육-충돌", nm, "주간" if s==0 else "야간"))
    if bad:
        print("[락 오류] 휴가/교육인데 x=1로 고정된 항목:")
        for d, t, nm, sh in bad:
            print(f"  - {d}일 {nm} {sh}")

    # 2) 날짜별 정원 초과(락만 기준)
    for d in range(1, days+1):
        for s, cap in [(0, day_size),(1, night_size)]:
            locked_ones = [(mi, members[mi]) for (mi,dd,ss),v in (hints or {}).get("locks",{}).get("x",{}).items()
                           if dd==d and ss==s and v==1]
            if len(locked_ones) > cap:
                print(f"[정원초과] {d}일 {'주간' if s==0 else '야간'}: 락된 1이 {len(locked_ones)}/{cap}")






# ------------------------
# 2) 머신러닝 및 AI 도입 (가중치 튜닝 & 과거 자료로 부터  스스로 학습.. 절대 손대지 말것)
# ------------------------
try:
    from skopt import gp_minimize
    from skopt.space import Integer
    _HAS_SKOPT = True
except Exception:
    _HAS_SKOPT = False

def solve_once(build_model_fn, weights: dict, hints: dict|None=None,
               time_limit=60, workers=8):
    model, H = build_model_fn(weights, hints)

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit
    solver.parameters.num_search_workers = workers
    status = solver.Solve(model)

    res = {"status": status, "feasible": status in (cp_model.FEASIBLE, cp_model.OPTIMAL)}
    try: res["obj"] = solver.ObjectiveValue()
    except Exception: res["obj"] = None

    try:
        x = H["x"]; members = H["members"]; DAYS = H["DAYS"]
        load_per = []
        for mi in range(len(members)):
            ld = sum(int(solver.BooleanValue(x[(mi,d,0)]) or solver.BooleanValue(x[(mi,d,1)]))
                     for d in range(1, DAYS+1))
            load_per.append(ld)
        res["load_std"] = float(np.std(load_per)) if load_per else 0.0
        res["load_maxmin_gap"] = (max(load_per)-min(load_per)) if load_per else 0
    except Exception:
        res["load_std"] = None; res["load_maxmin_gap"] = None

    try:
        schedule_builder = H.get("schedule_builder", None)
        if schedule_builder:
            res["schedule"] = schedule_builder(solver)
    except Exception:
        pass

    return res

def score_result(res: dict) -> float:
    if not res.get("feasible", False): return 1e9
    obj = res.get("obj", None)
    gap = res.get("load_maxmin_gap", 0) or 0
    std = res.get("load_std", 0.0) or 0.0
    if obj is None: return 1000 + 2*gap + std
    return float(obj) + 0.5*gap + 0.25*std

def tune_weights(build_model_fn, init_weights=None, n_calls=25, time_limit=60, workers=8):
    rng = random
    if init_weights is None:
        init_weights = {"alpha":10,"beta":1,"gamma":2,"delta":40,"ZETA":3,"LAMBDA_DN":400,"LAMBDA_RES":15}

    history = []

    def run_with(w):
        res = solve_once(build_model_fn, w, hints=None, time_limit=time_limit, workers=workers)
        sc = score_result(res)
        history.append({"weights": w, "score": sc, "feasible": res["feasible"], "obj": res.get("obj")})
        return sc

    if _HAS_SKOPT:
        space = [
            Integer(1,50,  name="alpha"),
            Integer(0,10,  name="beta"),
            Integer(0,10,  name="gamma"),
            Integer(5,60,  name="delta"),
            Integer(0,10,  name="ZETA"),
            Integer(0,1000,  name="LAMBDA_DN"),
            Integer(0,50,  name="LAMBDA_RES"),
        ]
        keys=["alpha","beta","gamma","delta","ZETA","LAMBDA_DN","LAMBDA_RES"]
        def f_to_minimize(x):
            w = {k:v for k,v in zip(keys, x)}
            return run_with(w)
        r = gp_minimize(f_to_minimize, space, n_calls=n_calls, n_initial_points=6)
        best_x = r.x
        best_w = {k:v for k,v in zip(keys, best_x)}
    else:
        warnings.warn("scikit-optimize 미탑재 → 랜덤 탐색으로 대체합니다.")
        best_w, best_s = None, float("inf")
        for _ in range(n_calls):
            w = {
                "alpha": rng.randint(1,50),
                "beta":  rng.randint(0,10),
                "gamma": rng.randint(0,10),
                "delta": rng.randint(5,60),
                "ZETA":  rng.randint(0,10),
                "LAMBDA_DN": rng.randint(0,1000),
                "LAMBDA_RES":rng.randint(0,50),
            }
            sc = run_with(w)
            if sc < best_s:
                best_s, best_w = sc, w

    hist_df = pd.DataFrame(history)
    return best_w, hist_df

# ---- 힌트용 간단 모델(과거 파일로부터 학습하는 모델) ----
def make_feature_table(members, YEAR, MONTH, DAYS, is_vac, is_k15):
    rows = []
    for mi, nm in enumerate(members):
        for d in range(1, DAYS+1):
            wkd = (pd.Timestamp(YEAR, MONTH, d).weekday() >= 5)
            for s in (0,1):
                rows.append({
                    "mi": mi, "name": nm, "day": d, "shift": s,
                    "is_vac": int(is_vac[(mi,d)]),
                    "is_k15": int(is_k15[mi]),
                    "is_weekend": int(wkd),
                    "dow": pd.Timestamp(YEAR, MONTH, d).weekday(),
                })
    return pd.DataFrame(rows)

def train_prob_model(past_csv_path: str):
    try:
        from sklearn.linear_model import LogisticRegression
        from sklearn.metrics import roc_auc_score
    except Exception:
        warnings.warn("scikit-learn 미탑재 → 힌트 학습 생략합니다.")
        return None

    if not os.path.exists(past_csv_path):
        return None
    df = pd.read_csv(past_csv_path, encoding="cp949")
    req = {"is_vac","is_k15","is_weekend","dow","shift","assigned"}
    if not req.issubset(df.columns):
        return None
    X = df[["is_vac","is_k15","is_weekend","dow","shift"]].values
    y = df["assigned"].values
    if y.sum() == 0: return None

    clf = LogisticRegression(max_iter=1000)
    clf.fit(X,y)
    try:
        auc = roc_auc_score(y, clf.predict_proba(X)[:,1])
        print(f"[hint] in-sample AUC={auc:.3f}")
    except Exception:
        pass
    return clf

def generate_x_hints(clf, members, YEAR, MONTH, DAYS, is_vac, is_k15,
                     DAY_SIZE=6, NIGHT_SIZE=6, prob_threshold=0.7):
    feats = make_feature_table(members, YEAR, MONTH, DAYS, is_vac, is_k15)
    if clf is None:

        x_hint = {}
        for d in range(1, DAYS+1):
            order = list(range(len(members)))
            random.shuffle(order)
            for s, cap in [(0, DAY_SIZE),(1,NIGHT_SIZE)]:
                cnt=0
                for mi in order:
                    if cnt>=cap: break
                    if is_vac[(mi,d)]==1: continue
                    if (mi,d,0) in x_hint or (mi,d,1) in x_hint: continue
                    x_hint[(mi,d,s)]=1; cnt+=1
        return x_hint

    X = feats[["is_vac","is_k15","is_weekend","dow","shift"]].values
    probs = clf.predict_proba(X)[:,1]
    feats["p"] = probs
    feats.loc[feats["is_vac"]==1, "p"] = -1.0

    x_hint = {}
    for d in range(1, DAYS+1):
        for s, cap in [(0, DAY_SIZE),(1, NIGHT_SIZE)]:
            sl = feats[(feats.day==d) & (feats.shift==s)].sort_values("p", ascending=False)
            picked=0
            for _, r in sl.iterrows():
                if picked>=cap: break
                if r["p"] < prob_threshold: break
                mi = int(r["mi"])
                if (mi,d,0) in x_hint or (mi,d,1) in x_hint: continue
                x_hint[(mi,d,s)] = 1
                picked += 1
    return x_hint



def make_training_hints(schedule: Dict[int, Dict[str, str]],
                        members: list,
                        training_ranges: list[tuple[str, int, int]],
                        year: int, month: int, days: int,
                        strict: bool = False):
    """
    핵심 변경점:
      - 교육 구간 날짜들(Tdays)에는 '누구도' 잠그지 않는다 (전원 언락).
      - 교육 대상자는 build_model_fn에서 is_vac=1 처리 → 자동 근무불가.
      - 교육 구간 밖은 기존 배정 고정(x와 y만). r는 굳이 잠글 필요 없음.
      - strict=False면 '교육 시작 전날'의 야간만 언락(전날-야간금지 충돌 완화).
    """
    name2mi = {normalize_name(n): i for i, n in enumerate(members)}
    # 교육 대상 mapping: mi -> (s,e)
    aff = {name2mi[normalize_name(n)]: (s, e)
           for (n, s, e) in training_ranges if normalize_name(n) in name2mi}

    # 교육 날짜들의 집합
    Tdays = set()
    for (s, e) in aff.values():
        s = max(1, s); e = min(days, e)
        for d in range(s, e+1):
            Tdays.add(d)

    locks = {"x": {}, "y": {}, "r": {}}

    # 교육 시작 전날들 (야간 언락용)
    start_minus1 = {(mi, max(1, s - 1)) for mi, (s, e) in aff.items()}

    def _to_shift_idx(sh):
        if sh is None: return None
        if isinstance(sh, int): return sh
        if isinstance(sh, str):
            u = sh.upper()
            if u in ("D","DAY","주간"): return 0
            if u in ("N","NIGHT","야간"): return 1
        return None

    for mi, m in enumerate(members):
        rng = aff.get(mi, None)
        for d in range(1, days + 1):
            tag = schedule[d].get(m, "")
            sh_raw, ell = _parse_shift_label(tag)  # "D/N/None" + ell 0/1/2/None
            sh = _to_shift_idx(sh_raw)

            # ▶ 교육 구간 날짜: 전원 언락 (아무 것도 잠그지 않음)
            if d in Tdays:
                continue

            # ▶ 교육 구간 밖: 기존 배정 고정 (x와 y만 고정, r는 굳이 잠그지 않음)
            if sh is None:
                # 휴가/예비/공란 등 → 근무는 0으로만 고정 (휴가는 build_model_fn의 is_vac로 강제됨)
                locks["x"][(mi, d, 0)] = 0
                locks["x"][(mi, d, 1)] = 0
            else:
                locks["x"][(mi, d, sh)] = 1
                locks["x"][(mi, d, 1 - sh)] = 0
                if ell is not None:
                    for e in range(3):
                        locks["y"][(mi, d, sh, e)] = int(e == ell)

    #  strict=False면: 교육 '시작 전날 야간' 언락
    if not strict:
        for mi, d in start_minus1:
            # 교육 전날이 교육 날짜 바로 앞이고, 그 전날 자체는 Tdays가 아님
            if d not in Tdays:
                locks["x"].pop((mi, d, 1), None)  # 야간 언락
                for e in range(3):
                    locks["y"].pop((mi, d, 1, e), None)

    return {"locks": locks, "training_ranges": training_ranges}




# ------------------------
# 3) CP-SAT
# ------------------------
def build_model_fn(weights: Dict[str,int], hints: Dict[str,Any]|None):
    members=load_members(FILE_SQUADS)
    existing_sched = (hints or {}).get("existing_schedule")

    if existing_sched:
      vac_set = vac_from_schedule(existing_sched, members, DAYS)
    else:
      vac_set = parse_vacation_sheet(FILE_VAC, YEAR, MONTH, DAYS)
    k15_set=load_k15_set(FILE_GUNS)

    M=len(members)
    try:
        rankbook = load_rankbook(FILE_RANK)
    except Exception:
        rankbook = {"role_by_name": {}, "leader_order": {}, "assistant_order": {}, "display_priority": {}}


    role_by_name_raw = rankbook.get("role_by_name", {})

    def _canon_role(v: str) -> str:
        t = str(v).strip().lower()
        if t in ("사수", "leader", "senior"): return "leader"
        if t in ("부사수", "assistant", "asso", "assist"): return "assistant"
        return ""


    role_by_name = {normalize_name(k): _canon_role(v) for k, v in role_by_name_raw.items()}
    is_assistant = {mi: int(role_by_name.get(normalize_name(members[mi])) == "assistant")
                    for mi in range(len(members))}
    # 사수/부사수 표준화 맵
    role_by_name_raw = rankbook.get("role_by_name", {})
    def _canon_role(v: str) -> str:
        t = str(v or "").strip().lower()
        if t in ("사수", "leader", "leaders", "senior"):
            return "leader"
        if t in ("부사수", "assistant", "assistants", "asso", "assist"):
            return "assistant"
        return ""



    training_ranges = (hints or {}).get("training_ranges", [])
    train_block = {
        (normalize_name(nm), d)
        for (nm, s, e) in training_ranges
        for d in range(max(1, s), min(DAYS, e) + 1)
    }
    vac_effect = vac_set | train_block



    is_vac={(mi,d):int((members[mi],d) in vac_effect) for mi in range(M) for d in range(1,DAYS+1)}
    is_k15={mi:int(members[mi] in k15_set) for mi in range(M)}



    weekend_days=[d for d in range(1,DAYS+1) if calendar.weekday(YEAR,MONTH,d)>=5]
    vac_days={m: sum((m,d) in vac_set for d in range(1, DAYS+1)) for m in members}
    max_vac= max(vac_days.values()) if vac_days else 0
    weight_by_mi = {mi: (1 + vac_days[members[mi]]) for mi in range(M)}

    model=cp_model.CpModel()
    x={(mi,d,s):model.NewBoolVar(f"x_{mi}_{d}_{s}") for mi in range(M) for d in range(1,DAYS+1) for s in (0,1)}
    r={(mi,d):model.NewBoolVar(f"r_{mi}_{d}") for mi in range(M) for d in range(1,DAYS+1)}
    y={(mi,d,s,ell):model.NewBoolVar(f"y_{mi}_{d}_{s}_{ell}")
       for mi in range(M) for d in range(1,DAYS+1) for s in (0,1) for ell in range(3)}

    # 정원
    for d in range(1,DAYS+1):
        model.Add(sum(x[(mi,d,0)] for mi in range(M))==DAY_SIZE)
        model.Add(sum(x[(mi,d,1)] for mi in range(M))==NIGHT_SIZE)
    # 개인-일자
    for mi in range(M):
        for d in range(1,DAYS+1):
            model.Add(x[(mi,d,0)]+x[(mi,d,1)]<=1)
    # 휴가 전날 야간금지조건
    for mi in range(M):
        for d in range(1,DAYS+1):
            if is_vac[(mi,d)]: model.Add(x[(mi,d,0)]==0); model.Add(x[(mi,d,1)]==0)
        for d in range(1,DAYS):
            if is_vac[(mi,d+1)]: model.Add(x[(mi,d,1)]==0)
    # 야→주 전환 금지조건
    for mi in range(M):
        for d in range(1,DAYS):
            model.Add(x[(mi,d,1)]+x[(mi,d+1,0)]<=1)

    # K15 사수 같은조 편성 금지 조건
    for d in range(1,DAYS+1):
        model.Add(sum(is_k15[mi]*x[(mi,d,0)] for mi in range(M))<=3)
        model.Add(sum(is_k15[mi]*x[(mi,d,1)] for mi in range(M))<=3)
        for s in (0,1):
            for ell in range(3):
                model.Add(sum(is_k15[mi]*y[(mi,d,s,ell)] for mi in range(M))<=1)

    for d in range(1, DAYS+1):
    # 주간: 부사수 2명까지 허용(= AA 가능)
        s = 0
        for ell in range(3):
            model.Add(sum(is_assistant[mi] * y[(mi, d, s, ell)] for mi in range(M)) <= 1) # 숫자통해서 부사수 조절할 수 있음

    # 야간: 부사수 동시 투입 금지(최대 1명)
        s = 1
        for ell in range(3):
            model.Add(sum(is_assistant[mi] * y[(mi, d, s, ell)] for mi in range(M)) <= 1)


    # 예비 개념 수치화
    for mi in range(M):
        for d in range(1,DAYS+1):
            model.Add(r[(mi,d)]+x[(mi,d,0)]+x[(mi,d,1)]+is_vac[(mi,d)]==1)


    # 예비 3연속 금지조건
    for mi in range(M):
        for d in range(1,DAYS-1):
           model.Add(r[(mi,d)]+r[(mi,d+1)]+r[(mi,d+2)]<=2)

    #야예야 금지 조건
    for mi in range(M):
         for d in range(2, DAYS):

             model.Add(x[(mi, d-1, 1)] + r[(mi, d)] + x[(mi, d+1, 1)] <= 2)


    for mi in range(M):
          for d in range(1,DAYS):
              if is_weekend(YEAR,MONTH,d) and is_weekend(YEAR,MONTH,d+1):
                  model.Add(r[(mi,d)]+r[(mi,d+1)]<=1)
          for d in range(2,DAYS-1):
            model.Add(x[(mi,d-1,1)]+r[(mi,d)]+r[(mi,d+1)]+x[(mi,d+2,1)]<=3)



    # ===========================================최대 근무 일수(수정가능)======================================================
    for mi in range(M):
        for s in (0,1):
            for d in range(1,DAYS-8+2):
                model.Add(sum(x[(mi,dd,s)] for dd in range(d,d+8))<=5) # <= 왼쪽에 있는 숫자(4) 수정하기
    #===========================================================================================================================

    starts, ends, t2_len2 = {}, {}, []
    t1_len1=[]
    t1_info = []
    t3_len3=[]
    for mi in range(M):
        for s in (0,1):
            for d in range(1,DAYS+1):
                st=model.NewBoolVar(f"st_{mi}_{d}_{s}")
                starts[(mi,d,s)]=st
                model.Add(st<=x[(mi,d,s)])
                if d>1:
                    model.Add(st<=1-x[(mi,d-1,s)])
                    model.Add(st>=x[(mi,d,s)]-x[(mi,d-1,s)])
                else:
                    model.Add(st==x[(mi,d,s)])



            for d in range(1,DAYS+1):
                e=model.NewBoolVar(f"end_{mi}_{d}_{s}")
                ends[(mi,d,s)]=e
                model.Add(e<=x[(mi,d,s)])
                if d<DAYS:
                    model.Add(e<=1-x[(mi,d+1,s)])
                    model.Add(e>=x[(mi,d,s)]-x[(mi,d+1,s)])
                else:
                    model.Add(e==x[(mi,d,s)])

            for d in range(1, DAYS+1):
                z1 = model.NewBoolVar(f"len1_{mi}_{d}_{s}")
                model.Add(z1 <= starts[(mi,d,s)])
                model.Add(z1 <= ends[(mi,d,s)])
                model.Add(z1 >= starts[(mi,d,s)] + ends[(mi,d,s)] - 1)
                t1_len1.append(z1)
                t1_info.append((z1, mi, s, d))


            for d in range(1,DAYS):
                z=model.NewBoolVar(f"len2_{mi}_{d}_{s}")
                model.Add(z<=starts[(mi,d,s)])
                model.Add(z<=ends[(mi,d+1,s)])
                model.Add(z>=starts[(mi,d,s)]+ends[(mi,d+1,s)]-1)
                t2_len2.append(z)

            for d in range(1, DAYS-1):
                z3 = model.NewBoolVar(f"len3_{mi}_{d}_{s}")
                model.Add(z3 <= starts[(mi,d,s)])
                model.Add(z3 <= ends[(mi,d+2,s)])
                model.Add(z3 >= starts[(mi,d,s)] + ends[(mi,d+2,s)] - 1)
                t3_len3.append(z3)





    def is_vac_day(mi, d):
        # 이미 build_model_fn 안에 is_vac 있음
        return is_vac[(mi, d)] == 1


    from collections import defaultdict
    per_mi_z1 = defaultdict(list)
    for (z1, mi, s, d) in t1_info:
        per_mi_z1[mi].append(z1)
    for mi, arr in per_mi_z1.items():
        model.Add(sum(arr) <= 1)






    # 조 인원배치

    for d in range(1,DAYS+1):
        for s in (0,1):
            for ell in range(3):
                model.Add(sum(y[(mi,d,s,ell)] for mi in range(M))==2)
            for mi in range(M):
                model.Add(sum(y[(mi,d,s,ell)] for ell in range(3))==x[(mi,d,s)])

    # --- (추가) 전월-당월 경계 연동 제약 (들여쓰기 주의) ---
    py, pm = prev_year_month(YEAR, MONTH)
    prev_json = f"rolling_state_{py}_{pm:02d}.json"
    rolling = load_rolling_context(prev_json)
    miss_bdry_rot = []

    unfair_coef = {mi: 1.0 for mi in range(M)}
    ETA = 1.0

    def _infer_last_shift_from_trw(trw: list) -> int|None:
        # d 내림차순: D/N이 처음 보이면 그걸 last_shift로
        for rec in sorted(trw or [], key=lambda r: r["d"], reverse=True):
            s = rec.get("shift")
            if s == "D": return 0
            if s == "N": return 1
        return None

    def _infer_last_label_from_trw(trw: list, target_shift_char: str) -> int|None:
        # d 내림차순: target 교대(D/N)에서 label이 처음 보이면 채택
        for rec in sorted(trw or [], key=lambda r: r["d"], reverse=True):
            if rec.get("shift") == target_shift_char and rec.get("label") is not None:
                return int(rec["label"])
        return None

    if rolling is not None:
        # 이름 매칭용
        name_to_mi = {normalize_name(nm): mi for mi, nm in enumerate(members)}
        prev_info = rolling.get("per_member", {})

        for raw_name, rec in prev_info.items():
            mi = name_to_mi.get(normalize_name(raw_name))
            if mi is None:
                continue


            last_shift = rec.get("last_shift", None)   # 0/1/None
            last_label = rec.get("last_label", None)   # 0/1/2/None
            trw = rec.get("trailing_week", []) or []
            L_MAX = 6  # 원하는 연속 상한(예: 최대 6일동안 동일한 근무 들어갈 수 있다.)
            def _carry_len(trw, s_char):
                cnt, last = 0, None
                for rrr in sorted(trw, key=lambda r: r["d"], reverse=True):
                    s = rrr.get("shift")
                    if s in ("D","N"):
                        if last is None:
                            last = s
                        if s == last == s_char:
                            cnt += 1
                        elif s != s_char:
                            break
                    else:
                        break
                return cnt

            kN = _carry_len(trw, "N")
            kD = _carry_len(trw, "D")

            if kN > 0:
                window = max(1, min(DAYS, L_MAX - kN + 1))
                model.Add(sum(x[(mi, d, 1)] for d in range(1, window + 1)) <= max(0, L_MAX - kN))
            if kD > 0:
                window = max(1, min(DAYS, L_MAX - kD + 1))
                model.Add(sum(x[(mi, d, 0)] for d in range(1, window + 1)) <= max(0, L_MAX - kD))

            # JSON이 틀렸더라도 trailing_week로 보정
            if last_shift not in (0, 1):
                last_shift = _infer_last_shift_from_trw(trw)
            if last_label not in (0, 1, 2) and last_shift in (0, 1):
                want = "D" if last_shift == 0 else "N"
                last_label = _infer_last_label_from_trw(trw, want)

            # === [경계 하드제약 #1] '야→주' 즉시 전환 금지 (야주 차단) ===
            if last_shift == 1:
                model.Add(x[(mi, 1, 0)] == 0)

            # === [경계 하드제약 #2] '야→예→야' 금지 (야예야 차단) ===
            if last_shift == 1 and DAYS >= 2:
                model.Add(x[(mi, 2, 1)] + r[(mi, 1)] <= 1)

            # === [라벨 회전] 같은 교대로 시작했다면 라벨은 반드시 회전 (조건부 강제) ===
            if last_shift in (0, 1) and last_label in (0, 1, 2):
                next_ell = (last_label + 1) % 3
                model.Add(y[(mi, 1, last_shift, next_ell)] >= x[(mi, 1, last_shift)])

                # (옵션) 회전 미스 카운트(소프트 페널티용) 유지
                mr = model.NewBoolVar(f"mr_{mi}")
                model.Add(mr >= x[(mi, 1, last_shift)] - y[(mi, 1, last_shift, next_ell)])
                miss_bdry_rot.append(mr)

            # (기존) prev_fairness → unfair_coef 계산 유지
            pf = rec.get("prev_fairness", None)
            if isinstance(pf, (int, float)):
                unfair = max(0.0, min(1.0, 1.0 - (pf / 100.0)))
                unfair_coef[mi] = 1.0 + ETA * unfair







    def _carry_lengths(trw):
        seq = sorted(trw, key=lambda r: r["d"], reverse=True)
        carry = {"D": 0, "N": 0, "R": 0}
        last = None
        for r in seq:
            s = r.get("shift")
            if s in ("D","N","R"):
                if last is None:
                    last = s; carry[s] += 1
                elif s == last:
                    carry[s] += 1
                else:
                    break
            else:
                break
        return carry

    def _last_label_for_shift(trw, target_shift):
        for rec in sorted(trw, key=lambda r: r["d"], reverse=True):
            if rec.get("shift") == target_shift and rec.get("label") is not None:
                return int(rec["label"])
        return None

    # 소프트 위반 페널티 가중치(필요시 조정)
    W_MIN_EXT = 5     # 최소길이 미충족 페널티
    W_MAX_CAP = 5     # 최대길이 초과 페널티
    W_R_CONT  = 5     # 예비 3연속 연속성 위반

    carry_viols = []

    if rolling is not None:
        name_to_mi = {normalize_name(nm): mi for mi, nm in enumerate(members)}
        prev_info = rolling.get("per_member", {})

        for raw_name, rec in prev_info.items():
            mi = name_to_mi.get(normalize_name(raw_name))
            if mi is None:
                continue

            trw = rec.get("trailing_week", []) or []
            carries = _carry_lengths(trw)  # {'D':kD, 'N':kN, 'R':kR}



            # 1) 11/1에 같은 교대를 배치하는 경우에만 라벨 회전 강제(조건부)
            for s_idx, s_char in [(0,"D"), (1,"N")]:
                last_lab = _last_label_for_shift(trw, s_char)
                if last_lab in (0,1,2):
                    next_ell = (last_lab + 1) % 3
                    # x(1,s)=1 일 때만 라벨 회전 강제
                    model.Add(y[(mi, 1, s_idx, next_ell)] >= x[(mi, 1, s_idx)])

            # 2) 동일 교대 최소/최대 길이 캐리 → x(1,s)=1 일 때만 조건부!!!
            for s_idx, s_char in [(0,"D"), (1,"N")]:
                k = int(carries.get(s_char, 0))
                m_need  = max(0, 3 - k)  # 최소 추가 일수
                M_allow = max(0, 5 - k)  # 허용 추가 상한

                # (a) 최소길이: 1일에 그 교대를 시작했다면 m_need-1일까지 end 금지
                if m_need >= 2:
                    # slack: 최소길이 위반
                    v_min = model.NewBoolVar(f"viol_minlen_{mi}_{s_idx}")
                    # x(1,s)=1 이고도 너무 일찍 끝났는지 감지
                    # "너무 일찍 끝났다"를 명시적으로 모델링하기 어려우므로, 간단히
                    # 1..(m_need-1)에서 해당 교대 합 >= (m_need-1) 이 되지 못하면 위반으로 본다.
                    model.Add(
                        sum(x[(mi, d, s_idx)] for d in range(1, min(DAYS, m_need) ))
                        + (1 - x[(mi,1,s_idx)])  # 1일에 안 넣으면 면책
                        >= (m_need - 1)
                    ).OnlyEnforceIf(v_min.Not())
                    carry_viols.append((v_min, W_MIN_EXT))

                # (b) 최대길이: 1..(M_allow+1) 창에서 같은 교대 합 ≤ M_allow
                # 1일에 그 교대를 쓰지 않으면 면책
                if M_allow < 5:
                    window = min(DAYS, M_allow + 1)
                    v_max = model.NewBoolVar(f"viol_maxlen_{mi}_{s_idx}")
                    # 위반이면 v_max=1 이 되도록:
                    # sum(x[1..window, s]) <= M_allow  (x(1,s)=1일 때만 의미)
                    # ⇔ sum(x[1..window, s]) - M_allow <= 0
                    # 완화: sum(x[1..window, s]) - M_allow - (1 - x(1,s))*window <= v_max*window
                    model.Add(
                        sum(x[(mi, d, s_idx)] for d in range(1, window + 1))
                        - M_allow
                        - (1 - x[(mi,1,s_idx)])*window
                        <= v_max * window
                    )
                    carry_viols.append((v_max, W_MAX_CAP))
                if trw:
                  # 전월 마지막 날(d가 가장 큰 날)의 상태
                  last_d = max(r["d"] for r in trw)
                  last_day_shift = next((r["shift"] for r in trw if r["d"] == last_d), None)

                  # '작업일(주/야)' 중 마지막으로 등장한 교대(D/N)를 찾되,
                  # 전월 마지막 날보다 앞쪽에서만 찾음
                  prev_work_shift = None
                  for rrec in sorted(trw, key=lambda r: r["d"], reverse=True):
                      if rrec["d"] < last_d and rrec.get("shift") in ("D","N"):
                          prev_work_shift = rrec["shift"]
                          break

                  # 패턴: (전월) ... N → [마지막날 R 또는 V] → (당월 1일) N  금지
                  if last_day_shift in ("R", "V") and prev_work_shift == "N":
                      # 당월 1일 '야간' 배치를 금지
                      model.Add(x[(mi, 1, 1)] == 0)



            kR = int(carries.get("R", 0))
            if kR > 0:
                need = max(0, 3 - kR)
                cap  = max(0, 2 - kR)
                if need >= 1:
                    end_d = min(DAYS, need)
                    v_r = model.NewBoolVar(f"viol_rescarry_{mi}")
                    # sum(r[1..end_d]) <= cap  (조건부로 완화: cap + v_r*end_d)
                    model.Add(sum(r[(mi, d)] for d in range(1, end_d + 1)) <= cap + v_r * end_d)
                    carry_viols.append((v_r, W_R_CONT))



    for mi in range(M):
        for s in (0, 1):  # 0=주간, 1=야간
            for d in range(1, DAYS):
                for ell in range(3):
                    next_ell = (ell + 1) % 3
                    model.Add(
                        y[(mi, d, s, ell)]
                        <= y[(mi, d+1, s, next_ell)] + (1 - x[(mi, d+1, s)])
                    )



    if hints and "locks" in hints and hints["locks"]:
        locks = hints["locks"]
        # x-locks
        for (mi,d,s), val in locks.get("x", {}).items():
            if (mi,d,s) in x:
                model.Add(x[(mi,d,s)] == int(val))
        # y-locks
        for (mi,d,s,ell), val in locks.get("y", {}).items():
            if (mi,d,s,ell) in y:
                model.Add(y[(mi,d,s,ell)] == int(val))
        # r-locks
        for (mi,d), val in locks.get("r", {}).items():
            if (mi,d) in r:
                model.Add(r[(mi,d)] == int(val))



    # 주말 예비 카운트(과도한 주말예비 독점방지)
    wknd_res_cnt=[]
    for mi in range(M):
        cnt=model.NewIntVar(0, len(weekend_days), f"wknd_res_{mi}")
        model.Add(cnt==sum(r[(mi,d)] for d in weekend_days))
        wknd_res_cnt.append(cnt)
    WKRmin = model.NewIntVar(0, len(weekend_days), "WKRmin")
    WKRmax = model.NewIntVar(0, len(weekend_days), "WKRmax")
    for c in wknd_res_cnt:
        model.Add(WKRmin <= c)
        model.Add(c <= WKRmax)



    # 근무량 균형
    loads=[model.NewIntVar(0,DAYS,f"load_{mi}") for mi in range(M)]
    for mi in range(M):
        model.Add(loads[mi]==sum(x[(mi,d,0)]+x[(mi,d,1)] for d in range(1,DAYS+1)))
    Lmin,Lmax=model.NewIntVar(0,DAYS,"Lmin"),model.NewIntVar(0,DAYS,"Lmax")
    for lm in loads: model.Add(Lmin<=lm); model.Add(lm<=Lmax)


    res_cnt = [model.NewIntVar(0, DAYS, f"res_cnt_{mi}") for mi in range(M)]
    for mi in range(M):
       model.Add(res_cnt[mi] == sum(r[(mi, d)] for d in range(1, DAYS+1)))

    Rmin = model.NewIntVar(0, DAYS, "Rmin")
    Rmax = model.NewIntVar(0, DAYS, "Rmax")
    for rc in res_cnt:
      model.Add(Rmin <= rc)
      model.Add(rc <= Rmax)


    switch_costs_by_mi = {mi: [] for mi in range(M)}
    for mi in range(M):
        for d in range(1, DAYS):
            sw = model.NewBoolVar(f"sw_{mi}_{d}")
            b1 = model.NewBoolVar(f"b1_{mi}_{d}")
            b2 = model.NewBoolVar(f"b2_{mi}_{d}")
            model.AddBoolAnd([x[(mi,d,0)], x[(mi,d+1,1)]]).OnlyEnforceIf(b1)
            model.AddBoolOr([x[(mi,d,0)].Not(), x[(mi,d+1,1)].Not()]).OnlyEnforceIf(b1.Not())
            model.AddBoolAnd([x[(mi,d,1)], x[(mi,d+1,0)]]).OnlyEnforceIf(b2)
            model.AddBoolOr([x[(mi,d,1)].Not(), x[(mi,d+1,0)].Not()]).OnlyEnforceIf(b2.Not())
            model.AddMaxEquality(sw, [b1, b2])
            switch_costs_by_mi[mi].append(sw)



    alpha=int(weights.get("alpha",10))
    beta =int(weights.get("beta",1))
    gamma=int(weights.get("gamma",2))
    delta=int(weights.get("delta",40))
    ZETA =int(weights.get("ZETA",3))
    LAMBDA_DN =int(weights.get("LAMBDA_DN",200))
    LAMBDA_RES =int(weights.get("LAMBDA_RES",15))
    REWARD_LEN2 = int(weights.get("delta", 40))
    LAMBDA_WKND_SPREAD = int(weights.get("LAMBDA_WKND_SPREAD", 150))

    day_cnt  = [model.NewIntVar(0, DAYS, f"day_cnt_{mi}")  for mi in range(M)]
    night_cnt= [model.NewIntVar(0, DAYS, f"night_cnt_{mi}")for mi in range(M)]
    diff_dn  = [model.NewIntVar(0, DAYS, f"diff_dn_{mi}")  for mi in range(M)]
    TAU_LEN1   = 40
    TAU_LEN2_R = 0
    TAU_LEN3_R = 6

    primary_obj=(
        alpha*(Lmax-Lmin)
        + beta *sum(unfair_coef[mi] * sum(switch_costs_by_mi[mi]) for mi in range(M))
        + gamma*sum(unfair_coef[mi] * weight_by_mi[mi] * wknd_res_cnt[mi] for mi in range(M))
        + TAU_LEN1   * sum(t1_len1)
        + LAMBDA_DN * sum(unfair_coef[mi] * diff_dn[mi] for mi in range(M))
        + LAMBDA_RES* (Rmax-Rmin)
        + LAMBDA_WKND_SPREAD * (WKRmax - WKRmin)
        - TAU_LEN3_R * sum(t3_len3)
    )



    for mi in range(M):
      model.Add(day_cnt[mi]   == sum(x[(mi, d, 0)] for d in range(1, DAYS+1)))
      model.Add(night_cnt[mi] == sum(x[(mi, d, 1)] for d in range(1, DAYS+1)))

      tmp = model.NewIntVar(-DAYS, DAYS, f"tmp_nd_{mi}")
      model.Add(tmp == night_cnt[mi] - day_cnt[mi])
      model.Add(diff_dn[mi] >= tmp)
      model.Add(diff_dn[mi] >= -tmp)
      #============================================================================================
      DN_CAP = 4   #ㅣ주간-야간ㅣ 을 의미한다. 원한다면 수정가능..!!!!
      #============================================================================================
      for mi in range(M):
         model.Add(diff_dn[mi] <= DN_CAP)

         tie_dn_bias = sum(diff_dn)


    tie_consistency = sum(
    (mi + 1) * sum(x[(mi, d, s)] for d in range(1, DAYS+1) for s in (0, 1))
    for mi in range(M)
)

    BIG_M = 10_000_000
    MID_M = 5_000
    ROT_BDRY_W = 20


    if carry_viols:
       carry_penalty = sum(w * v for (v, w) in carry_viols)
    else:
       carry_penalty = 0

    model.Minimize(BIG_M * primary_obj + MID_M * tie_dn_bias + tie_consistency+ ROT_BDRY_W * sum(miss_bdry_rot)+ carry_penalty)


    if hints:
        if "x" in hints:
            for k,v in hints["x"].items():
                if k in x: model.AddHint(x[k], int(v))
        if "y" in hints:
            for k,v in hints["y"].items():
                if k in y: model.AddHint(y[k], int(v))
        if "r" in hints:
            for k,v in hints["r"].items():
                if k in r: model.AddHint(r[k], int(v))

    def schedule_builder(solver):
        LABELS = ['C','B','A']
        schedule = {d: {} for d in range(1, DAYS+1)}
        for d in range(1, DAYS+1):
            for mi, m in enumerate(members):
                if is_vac[(mi, d)] == 1:
                    schedule[d][m] = "휴가"; continue
                is_day   = solver.BooleanValue(x[(mi, d, 0)])
                is_night = solver.BooleanValue(x[(mi, d, 1)])
                if is_day:
                    ell = next((e for e in range(3) if solver.BooleanValue(y[(mi, d, 0, e)])), None)
                    lab = LABELS[ell] if ell is not None else "?"
                    schedule[d][m] = f"주간 {lab}조"
                elif is_night:
                    ell = next((e for e in range(3) if solver.BooleanValue(y[(mi, d, 1, e)])), None)
                    lab = LABELS[ell] if ell is not None else "?"
                    schedule[d][m] = f"야간 {lab}조"
                else:
                    schedule[d][m] = "예비"
        return schedule

    var_handles = {
        "x": x, "y": y, "r": r,
        "members": members, "DAYS": DAYS,
        "schedule_builder": schedule_builder
    }
    return model, var_handles


# ------------------------
# 4) Print .xlsx
# ------------------------
def export_schedule_to_excel(schedule: Dict[int, Dict[str,str]], members: list, out_path: str,
                             year:int, month:int, days:int, rankbook: Dict[str,dict]|None=None):


    if rankbook is None:
      try:
        rankbook = load_rankbook(FILE_RANK)
      except Exception:
            rankbook = {"role_by_name": {}, "leader_order": {}, "assistant_order": {}, "display_priority": {}}
    role_by_name_raw = rankbook["role_by_name"]

    def _canon_role(v: str) -> str:
        t = str(v).strip().lower()
        if t in ("사수", "leader", "senior"):    return "leader"
        if t in ("부사수", "assistant"):         return "assistant"
        return ""

    # 키(이름)는 normalize, 값(역할)은 leader/assistant로 표준화
    role_by_name = {normalize_name(k): _canon_role(v) for k, v in role_by_name_raw.items()}

    # 순서/표시 우선순위 맵도 키를 normalize해서 맞춰줌
    leader_order     = {normalize_name(k): v for k, v in rankbook.get("leader_order", {}).items()}
    assistant_order  = {normalize_name(k): v for k, v in rankbook.get("assistant_order", {}).items()}
    display_priority = {normalize_name(k): v for k, v in rankbook.get("display_priority", {}).items()}


    wb = Workbook()
    ws_month = wb.active
    ws_month.title = "월간"
    ws_month.append(["이름"] + [str(d) for d in range(1, days+1)])
    ws_month.column_dimensions["A"].width = 9
    for d in range(1, days+1):
      col_letter=ws_month.cell(row=1, column=d+1).column_letter
      ws_month.column_dimensions[col_letter].width = 12


    for m in members:
        row = {"이름": m}
        for d in range(1, days+1):
            row[str(d)] = schedule[d].get(m, "")
        ws_month.append([row[k] for k in ["이름"]+[str(d) for d in range(1,days+1)]])

    summary_start_col = 1 + days + 1  # 1:이름, 2~(1+days): 날짜열
    ws_month.cell(row=1, column=summary_start_col + 0, value="주간합")
    ws_month.cell(row=1, column=summary_start_col + 1, value="야간합")
    ws_month.cell(row=1, column=summary_start_col + 2, value="예비합")
    ws_month.cell(row=1, column=summary_start_col + 3, value="휴가합")
    ws_month.cell(row=1, column=summary_start_col + 4, value="주/야 비율")
    ws_month.cell(row=1, column=summary_start_col + 5, value="균형 차이")
    ws_month.cell(row=1, column=summary_start_col + 6, value="총근무합")
    ws_month.cell(row=1, column=summary_start_col + 7, value="주말예비합")
    ws_month.cell(row=1, column=summary_start_col + 8, value="공정성 점수")


    # ==== 요약/공정성: 주/야 비율·균형 차이 열 제거 & 오른쪽 열 당김 ====
    start_col_letter = get_column_letter(2)          # 첫 날짜 열(B)
    end_col_letter   = get_column_letter(1 + days)   # 마지막 날짜 열
    weekend_cols = [get_column_letter(1 + d) for d in range(1, days+1)
                    if calendar.weekday(year, month, d) >= 5]

    max_row = ws_month.max_row
    summary_start_col = 1 + days + 1  # 이름열 + 날짜열 다음 위치

    # 헤더: 비율/균형 제거 → 총근무합/주말예비합/공정성 점수 왼쪽으로 이동
    headers = ["주간합","야간합","예비합","휴가합","총근무합","주말예비합","공정성 점수"]
    for i, h in enumerate(headers):
        ws_month.cell(row=1, column=summary_start_col + i, value=h)

    # COUNTIF
    for r in range(2, max_row + 1):
        row_range = f"{start_col_letter}{r}:{end_col_letter}{r}"

        ws_month.cell(row=r, column=summary_start_col + 0).value = f'=COUNTIF({row_range},"*주간*")'
        ws_month.cell(row=r, column=summary_start_col + 1).value = f'=COUNTIF({row_range},"*야간*")'
        ws_month.cell(row=r, column=summary_start_col + 2).value = f'=COUNTIF({row_range},"예비")'
        ws_month.cell(row=r, column=summary_start_col + 3).value = f'=COUNTIF({row_range},"휴가")'

        day_addr   = ws_month.cell(row=r, column=summary_start_col + 0).coordinate
        night_addr = ws_month.cell(row=r, column=summary_start_col + 1).coordinate
        ws_month.cell(row=r, column=summary_start_col + 4).value = f'={day_addr}+{night_addr}'


        wk_res_cell = ws_month.cell(row=r, column=summary_start_col + 5)
        if weekend_cols:
            cntifs = "+".join([f'COUNTIF({col}{r},"예비")' for col in weekend_cols])
            wk_res_cell.value = f'={cntifs}'
        else:
            wk_res_cell.value = 0


    W_RATIO = 0.10 #주야비율.. 주간 많을수록 +
    W_RES   = 0.60 #총 예비일수 많을수록 +
    W_LOW   = 0.30 #총 근무일수 많을수록 -

    total_col_letter = get_column_letter(summary_start_col + 4)   # 총근무합
    res_col_letter   = get_column_letter(summary_start_col + 2)   # 예비합
    max_total_range  = f'${total_col_letter}$2:${total_col_letter}${max_row}'
    max_res_range    = f'${res_col_letter}$2:${res_col_letter}${max_row}'

    for r in range(2, max_row + 1):
        day_addr   = ws_month.cell(row=r, column=summary_start_col + 0).coordinate  # 주간합
        night_addr = ws_month.cell(row=r, column=summary_start_col + 1).coordinate  # 야간합
        res_addr   = ws_month.cell(row=r, column=summary_start_col + 2).coordinate  # 예비합
        total_addr = ws_month.cell(row=r, column=summary_start_col + 4).coordinate  # 총근무합

        # 비율/균형 열은 생성하지 않고, 공정성 점수 수식 안에서 직접 계산
        ratio_score   = f'IF(({day_addr}+{night_addr})=0,0.5,{day_addr}/({day_addr}+{night_addr}))'
        reserve_score = f'IF(MAX({max_res_range})=0,0,{res_addr}/MAX({max_res_range}))'
        lowload_score = f'IF(MAX({max_total_range})=0,1,1-({total_addr}/MAX({max_total_range})))'

        fair_formula  = (
            f'=ROUND(100*('
            f'{W_RATIO}*({ratio_score}) + {W_RES}*({reserve_score}) + {W_LOW}*({lowload_score})'
            f'),0)'
        )
        ws_month.cell(row=r, column=summary_start_col + 6).value = fair_formula  # (변경) 공정성 점수 위치


    try:
      wb.calculation_properties.fullCalcOnLoad = True
    except AttributeError:
      try:
          wb.calcPr.fullCalcOnLoad = True
      except Exception:
          pass


    first_data_row = 2
    last_data_row  = ws_month.max_row
    name_col_idx   = 1
    fair_col_idx   = summary_start_col + 8

    BLUE = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for d in range(1, days+1):
        if calendar.weekday(year, month, d) >= 5:
            col = d + 1
            for row in ws_month.iter_rows(min_row=1, max_row=ws_month.max_row, min_col=col, max_col=col):
                for cell in row:
                    cell.fill = BLUE



    def _pick_pair(role_cands, order_map):
      # order_map(숫자 작을수록 상위)로 정렬해 1명 고름
        return sorted(role_cands, key=lambda nm: order_map.get(nm, 10**6))[0] if role_cands else None


    def _is_leader(nm: str) -> bool:
        return role_by_name.get(normalize_name(nm)) == "leader"

    def _is_assistant(nm: str) -> bool:
        return role_by_name.get(normalize_name(nm)) == "assistant"

    def _sort_leaders(nms):    return sorted(nms, key=lambda nm: leader_order.get(normalize_name(nm), 10**6))
    def _sort_assistants(nms): return sorted(nms, key=lambda nm: assistant_order.get(normalize_name(nm), 10**6))
    def _sort_display(nms):    return sorted(nms, key=lambda nm: display_priority.get(normalize_name(nm), 10**6))

    def names_by(shift_label: str, team_letter: str, day_dict: dict) -> str:
        cands = [nm for nm, tag in day_dict.items()
                if tag.startswith(shift_label) and f"{team_letter}조" in tag]
        if not cands:
            return ""

        leaders    = _sort_leaders([nm for nm in cands if _is_leader(nm)])
        assistants = _sort_assistants([nm for nm in cands if _is_assistant(nm)])

        if leaders and assistants:                     # 사수/부사수 1명씩 → 사수/부사수
            return f"{leaders[0]}/{assistants[0]}"
        if len(leaders) >= 2:                          # 사수/사수 → 순위 상위 2명
            return f"{leaders[0]}/{leaders[1]}"


        # 역할이 비어있거나 섞였는데 1~2명인 경우: 표시우선순위로 좌/우, 가능하면 사수 좌/부사수 우로 보정
        cands_sorted = _sort_display(cands)
        if len(cands_sorted) == 1:
            return cands_sorted[0]
        left, right = cands_sorted[0], cands_sorted[1]
        if _is_assistant(left) and _is_leader(right):
            left, right = right, left
        return f"{left}/{right}"





    def reserve_groups(day_dict):
        res = [nm for nm, tag in day_dict.items() if tag == "예비"]
        buckets = []
        i = 0
        for lab in ["A","B","C","D","E"]:
            a = res[i]   if i   < len(res) else ""
            b = res[i+1] if i+1 < len(res) else ""
            pair = [a, b]
            pair.sort(key=lambda nm: display_priority.get(nm, 10**6))
            left, right = pair
            buckets.append((lab, f"{left}/{right}" if right else left))
            i += 2
        return buckets



    for d in range(1, days+1):
        ws = wb.create_sheet(title=f"{month:02d}-{d:02d}")
        ws.append([f"{year}년 {month}월 {d}일 경계작전 지침(요약)"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.append(["", "주간 분대 편성", "", "", "야간 분대 편성", ""])
        ws.append(["구분", "관등성명1/2", "", "구분", "관등성명1/2"])

        dayA = names_by("주간","A", schedule[d])
        dayB = names_by("주간","B", schedule[d])
        dayC = names_by("주간","C", schedule[d])
        ngtA = names_by("야간","A", schedule[d])
        ngtB = names_by("야간","B", schedule[d])
        ngtC = names_by("야간","C", schedule[d])

        ws.append(["A", dayA, "", "A", ngtA])
        ws.append(["B", dayB, "", "B", ngtB])
        ws.append(["C", dayC, "", "C", ngtC])

        ws.append([])
        ws.append(["예비분대"]); ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
        ws.append(["구분", "관등성명1/2"])
        for lab, joined in reserve_groups(schedule[d]):
            ws.append([lab, joined])

        for col in ["A","B","C","D","E","F"]:
            ws.column_dimensions[col].width = 16


    wb.save(out_path)
    print("저장 완료 →", out_path)

# ------------------------
# 5) Main Flow
# ------------------------
def main():
    # 과거의 데이터 학습.(아무 공정표나 상관없다.공정표의 흐름을 보는것.. clf ""안에는 반드시 엑셀을 csv로 변환한 파일의 제목을 넣어야한다.)
    clf = train_prob_model("ml.csv")
    members = load_members(FILE_SQUADS)
    vac_set = parse_vacation_sheet(FILE_VAC, YEAR, MONTH, DAYS)
    k15_set = load_k15_set(FILE_GUNS)
    M = len(members)
    is_vac = {(mi,d): int((members[mi], d) in vac_set) for mi in range(M) for d in range(1, DAYS+1)}
    is_k15 = {mi: int(members[mi] in k15_set) for mi in range(M)}


    x_hints = generate_x_hints(
        clf, members, YEAR, MONTH, DAYS, is_vac, is_k15,
        DAY_SIZE=DAY_SIZE, NIGHT_SIZE=NIGHT_SIZE, prob_threshold=0.60
    )
    hints = {"x": x_hints}
    best_w, hist = tune_weights(build_model_fn, n_calls=12, time_limit=30, workers=8)
    print("[best weights]", best_w)
    best_w["LAMBDA_WKND_SPREAD"] = 400
    try:
        hist.to_csv("weight_search_log.csv", index=False)
    except Exception:
        pass


    res = solve_once(build_model_fn, best_w, hints=hints, time_limit=25, workers=8)
    print("status:", res["status"], "feasible:", res["feasible"], "obj:", res.get("obj"))

    if res.get("feasible", False) and "schedule" in res:
        try:
            rankbook = load_rankbook(FILE_RANK)
        except Exception:
            rankbook = None
        export_schedule_to_excel(res["schedule"], members, OUT_PATH, YEAR, MONTH, DAYS, rankbook)
        save_rolling_context_from_schedule(
            f"rolling_state_{YEAR}_{MONTH:02d}.json",
            res["schedule"], members, YEAR, MONTH, DAYS
        )

    else:
        print("해를 찾지 못했습니다. (하드 제약 충돌 가능)")

if __name__ == "__main__":
    RUN_BUILD =True
    RUN_REPAIR =False

    if RUN_BUILD:
        main()

    if RUN_REPAIR:
        members = load_members(FILE_SQUADS)

        #existing에 파일명 직접 지정
        existing ="수정필요.xlsx"
        sch = load_schedule_from_excel(existing, members, DAYS)

        #교육(근무불가) 구간 지정... 모든 조건 다 만족..
        training = [("유근서",3,3)]


        hints = make_training_hints(sch, members, training, YEAR, MONTH, DAYS, strict=False)
        hints["existing_schedule"] = sch

        w = {"alpha":10, "beta":1, "gamma":2, "delta":40, "ZETA":3, "LAMBDA_DN":25, "LAMBDA_RES":15}
        res_fix = solve_once(build_model_fn, w, hints=hints, time_limit=40, workers=8)

        if res_fix.get("feasible", False) and "schedule" in res_fix:
            try:
                rankbook = load_rankbook(FILE_RANK)
            except Exception:
                rankbook = None
            export_schedule_to_excel(
                res_fix["schedule"], members,
                f"{YEAR}년_{MONTH:02d}월_공정표_repair.xlsx",
                YEAR, MONTH, DAYS, rankbook
            )
            print("✅ 수정본 저장 완료")
        else:
            print("❌ 수정 실패: 이 기간에 헤당인원이 교육/외진을 간다면 그 인원 대신 근무 투입을 하는 동시에 규칙을 만족하는 경우의 수는 존재하지 않습니다. 교육이라면 다른 인원을 선택하는걸 추천드립니다.")

