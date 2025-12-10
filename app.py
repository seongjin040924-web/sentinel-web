import os
import io

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

import sentinel_core  # 네 메인 근무표 코드


# ---------------------------------------------------
# 0) 공통 설정 & 유틸
# ---------------------------------------------------
st.set_page_config(
    page_title="Sentinel AIP-lite",
    layout="wide",
    page_icon="🛡️",
)

# 캐치테이블 + Palantir 느낌의 간단한 CSS
st.markdown(
    """
    <style>
    .main {
        background: #0b1724;
        color: #f5f7fb;
        font-family: -apple-system, BlinkMacSystemFont, "SF Pro Text", system-ui, sans-serif;
    }
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #020617 0%, #020617 60%, #020617 100%);
        color: #e5e7eb;
    }
    .stButton>button {
        border-radius: 999px;
        padding: 0.45rem 1.2rem;
        border: 1px solid #1e293b;
        background: linear-gradient(135deg, #0ea5e9, #22c55e);
        color: white;
        font-weight: 600;
    }
    .card {
        border-radius: 18px;
        padding: 18px 18px 14px 18px;
        background: rgba(15, 23, 42, 0.9);
        border: 1px solid rgba(148, 163, 184, 0.35);
        box-shadow: 0 18px 50px rgba(15, 23, 42, 0.7);
    }
    .card-title {
        font-size: 0.95rem;
        color: #9ca3af;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 0.1rem;
    }
    .card-value {
        font-size: 1.8rem;
        font-weight: 700;
        color: #e5e7eb;
    }
    .pill {
        display: inline-flex;
        align-items: center;
        padding: 4px 10px;
        border-radius: 999px;
        border: 1px solid #1f2937;
        background: rgba(15,23,42,0.9);
        font-size: 0.75rem;
        color: #9ca3af;
    }
    .tag {
        display: inline-flex;
        padding: 2px 8px;
        border-radius: 999px;
        font-size: 0.7rem;
        border: 1px solid #1f2937;
        color: #6b7280;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

OUT_PATH = sentinel_core.OUT_PATH


def load_monthly_df():
    """월간 시트를 DataFrame으로 로드."""
    if not os.path.exists(OUT_PATH):
        return None
    try:
        df = pd.read_excel(OUT_PATH, sheet_name="월간")
        return df
    except Exception as e:
        st.error(f"엑셀 로드 실패: {e}")
        return None


def detect_day_columns(df: pd.DataFrame):
    """1,2,3,... 형태의 날짜 열만 골라서 리스트로 반환."""
    day_cols = []
    for c in df.columns:
        if isinstance(c, int):
            day_cols.append(c)
        elif isinstance(c, str) and c.isdigit():
            day_cols.append(c)
    return day_cols


def compute_member_stats(df: pd.DataFrame):
    """
    일자별 텍스트를 이용해 간단한 통계 계산.
    (주간/야간/예비/휴가 횟수)
    """
    if "이름" not in df.columns:
        st.error("월간 시트에 '이름' 열이 없습니다.")
        return None

    day_cols = detect_day_columns(df)
    rows = []
    for _, row in df.iterrows():
        name = str(row["이름"])
        tags = [str(row[c]) for c in day_cols]
        day_cnt = sum("주간" in t for t in tags)
        night_cnt = sum("야간" in t for t in tags)
        reserve_cnt = sum(t == "예비" for t in tags)
        vac_cnt = sum(t == "휴가" for t in tags)
        total = day_cnt + night_cnt

        rows.append(
            {
                "이름": name,
                "주간": day_cnt,
                "야간": night_cnt,
                "총근무": total,
                "예비": reserve_cnt,
                "휴가": vac_cnt,
            }
        )

    return pd.DataFrame(rows)


def apply_manual_edits_to_excel(edited_df: pd.DataFrame):
    """
    st.data_editor로 수정한 월간표를 엑셀에 반영.
    날짜 열(1~DAYS)만 업데이트하고 오른쪽 요약/공정성 수식은 유지.
    """
    if not os.path.exists(OUT_PATH):
        st.error("기존 공정표 파일이 없습니다.")
        return None

    wb = load_workbook(OUT_PATH, data_only=False)
    if "월간" not in wb.sheetnames:
        st.error("엑셀에 '월간' 시트를 찾을 수 없습니다.")
        return None

    ws = wb["월간"]
    day_cols = detect_day_columns(edited_df)

    # 엑셀 상에서: 1행 = 헤더, 2행부터 데이터
    for i, (_, row) in enumerate(edited_df.iterrows()):
        excel_row = i + 2
        for c in day_cols:
            # 열 인덱스: 이름(A)=1, 1일=2, ... 이므로 날짜열 = 1 + int(day)
            col_idx = 1 + int(c)
            ws.cell(row=excel_row, column=col_idx).value = row[c]

    # 기존 파일은 살려두고 수정본 별도 저장
    new_path = OUT_PATH.replace(".xlsx", "_수정본.xlsx")
    wb.save(new_path)
    return new_path


def rebuild_with_newcomer(new_name: str, squad_label: str):
    """
    신병을 분대편성표에 추가한 뒤, 해당 편성으로 다시 공정표 생성.
    squad_label 예: '1분대', '2분대', ...
    """
    squads_path = sentinel_core.FILE_SQUADS
    if not os.path.exists(squads_path):
        st.error(f"분대편성표 파일을 찾을 수 없습니다: {squads_path}")
        return None

    df = pd.read_excel(squads_path)

    # '분대', '이름' 열을 가진 포맷이라고 가정
    if "분대" not in df.columns or "이름" not in df.columns:
        st.error("분대편성표에 '분대'와 '이름' 열이 필요합니다.")
        return None

    # 이미 존재하는 이름인지 체크
    if str(new_name) in df["이름"].astype(str).tolist():
        st.warning("이미 분대편성표에 존재하는 이름입니다. 그대로 재생성만 수행합니다.")

    else:
        df = pd.concat(
            [
                df,
                pd.DataFrame(
                    [
                        {
                            "분대": squad_label,
                            "이름": new_name,
                        }
                    ]
                ),
            ],
            ignore_index=True,
        )

    tmp_squads = squads_path.replace(".xlsx", "_신병반영.xlsx")
    df.to_excel(tmp_squads, index=False)

    # sentinel_core에서 사용하는 FILE_SQUADS를 잠깐 변경
    old_squads = sentinel_core.FILE_SQUADS
    sentinel_core.FILE_SQUADS = tmp_squads

    try:
        sentinel_core.main()
    finally:
        sentinel_core.FILE_SQUADS = old_squads

    return sentinel_core.OUT_PATH


# ---------------------------------------------------
# 1) 사이드바 & 상단 헤더
# ---------------------------------------------------
with st.sidebar:
    st.markdown("### 🛡 Sentinel AIP-lite")
    st.caption("2소대 경계작전 공정표 · Palantir 스타일 대시보드")

    page = st.radio(
        "페이지 선택",
        [
            "대시보드",
            "근무표 생성",
            "개별 통계",
            "신병 투입 / 팀 재배치",
            "근무표 수동 수정",
            "디자인 Mock-up",
        ],
    )

    st.markdown("---")
    st.markdown(
        """
        <div class="tag">v0.1 • prototype</div>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
      <div class="pill">Ops / Guard · Sentinel</div>
    </div>
    <h1 style="margin:0;font-size:2.1rem;font-weight:700;">
      Sentinel Guard Planner
    </h1>
    <p style="margin-top:4px;color:#9ca3af;font-size:0.9rem;">
      Palantir AIP 스타일로 2소대 경계작전 공정표를 설계, 재배치, 모니터링합니다.
    </p>
    """,
    unsafe_allow_html=True,
)

st.write("")


# ---------------------------------------------------
# 2) 대시보드
# ---------------------------------------------------
if page == "대시보드":
    df_month = load_monthly_df()
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">현재 공정표 파일</div>', unsafe_allow_html=True)
        if os.path.exists(OUT_PATH):
            st.markdown(f'<div class="card-value">✅ 생성됨</div>', unsafe_allow_html=True)
            st.caption(os.path.basename(OUT_PATH))
        else:
            st.markdown(f'<div class="card-value">⚠ 미생성</div>', unsafe_allow_html=True)
            st.caption("먼저 [근무표 생성] 탭에서 생성하세요.")
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">현재 월</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="card-value">{sentinel_core.YEAR}년 {sentinel_core.MONTH}월</div>',
            unsafe_allow_html=True,
        )
        st.caption(f"총 일수: {sentinel_core.DAYS}일")
        st.markdown('</div>', unsafe_allow_html=True)

    with col3:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">총 인원</div>', unsafe_allow_html=True)
        try:
            members = sentinel_core.load_members(sentinel_core.FILE_SQUADS)
            st.markdown(f'<div class="card-value">{len(members)}명</div>', unsafe_allow_html=True)
            st.caption("분대편성표 기준")
        except Exception:
            st.markdown('<div class="card-value">-</div>', unsafe_allow_html=True)
            st.caption("분대편성표 로드 실패")
        st.markdown('</div>', unsafe_allow_html=True)

    with col4:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<div class="card-title">ML 힌트 사용</div>', unsafe_allow_html=True)
        st.markdown('<div class="card-value">ON</div>', unsafe_allow_html=True)
        st.caption("ml.csv 기반 확률 힌트 + 가중치 자동 튜닝")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")

    st.subheader("요약 뷰")

    if df_month is not None:
        st.caption("현재 월간 공정표 일부 미리보기 (상위 15명)")
        st.dataframe(df_month.head(15), use_container_width=True)
    else:
        st.info("아직 공정표가 없어요. 왼쪽에서 [근무표 생성] 탭으로 이동해서 먼저 생성하세요.")


# ---------------------------------------------------
# 3) 근무표 생성
# ---------------------------------------------------
elif page == "근무표 생성":
    st.subheader("근무표 생성 / 재생성")

    with st.expander("현재 설정 확인", expanded=True):
        st.write(f"📅 대상: **{sentinel_core.YEAR}년 {sentinel_core.MONTH}월**")
        st.write(f"📁 분대편성표: `{os.path.basename(sentinel_core.FILE_SQUADS)}`")
        st.write(f"📁 휴가 파일: `{os.path.basename(sentinel_core.FILE_VAC)}`")
        st.write(f"📁 총기 파일: `{os.path.basename(sentinel_core.FILE_GUNS)}`")
        st.write(f"📁 짬표: `{os.path.basename(sentinel_core.FILE_RANK)}`")
        st.write(f"📤 출력 파일: `{os.path.basename(OUT_PATH)}`")

    st.warning(
        "⚠ 실행 시 OR-Tools + ML 가중치 튜닝이 돌아서, Colab 기준으로 시간이 꽤 걸릴 수 있습니다."
    )

    if st.button("🔁 공정표 생성 / 재생성", use_container_width=True):
        with st.spinner("CP-SAT 모델로 최적 공정표 계산 중... ⏳"):
            sentinel_core.main()
        st.success("✅ 공정표 생성 완료!")

    df_month = load_monthly_df()
    if df_month is not None:
        st.markdown("### 생성된 월간 공정표 (상위 20명)")
        st.dataframe(df_month.head(20), use_container_width=True)

        # 다운로드 버튼 (엑셀 그대로)
        with open(OUT_PATH, "rb") as f:
            st.download_button(
                "엑셀 파일 다운로드",
                data=f,
                file_name=os.path.basename(OUT_PATH),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ---------------------------------------------------
# 4) 개별 통계
# ---------------------------------------------------
elif page == "개별 통계":
    st.subheader("개별 병사 근무 통계 / 그래프")

    df_month = load_monthly_df()
    if df_month is None:
        st.info("아직 생성된 공정표가 없습니다. 먼저 [근무표 생성] 탭에서 생성하세요.")
    else:
        stats_df = compute_member_stats(df_month)
        if stats_df is None:
            st.stop()

        col_left, col_right = st.columns([1.5, 2])

        with col_left:
            st.markdown("#### 인원 리스트")
            st.dataframe(stats_df, use_container_width=True, height=400)

            member_names = stats_df["이름"].tolist()
            target = st.selectbox("📌 상세 분석할 병사 선택", member_names)

        with col_right:
            st.markdown("#### 선택 인원 근무 분포")

            row = stats_df[stats_df["이름"] == target].iloc[0]
            chart_df = pd.DataFrame(
                {
                    "구분": ["주간", "야간", "예비", "휴가"],
                    "횟수": [row["주간"], row["야간"], row["예비"], row["휴가"]],
                }
            ).set_index("구분")

            st.bar_chart(chart_df)

            st.caption(
                f"총 근무: **{int(row['총근무'])}일**, 예비: **{int(row['예비'])}일**, 휴가: **{int(row['휴가'])}일**"
            )


# ---------------------------------------------------
# 5) 신병 투입 / 팀 재배치
# ---------------------------------------------------
elif page == "신병 투입 / 팀 재배치":
    st.subheader("신병 투입 시 자동 팀 재배치 (v0.1 Prototype)")

    st.write(
        "분대편성표에 신병을 추가한 뒤, 같은 규칙으로 **새 공정표를 재생성**합니다. "
        "아직은 세부 팀(A/B/C)까지 직접 조정하는 단계는 아니고, CP-SAT가 전체를 다시 최적화하는 방식입니다."
    )

    col1, col2 = st.columns(2)
    with col1:
        new_name = st.text_input("신병 이름", placeholder="예: 홍길동")
    with col2:
        squad_label = st.selectbox("배정 분대", ["1분대", "2분대", "3분대"])

    if st.button("🧩 신병 반영 & 공정표 재생성", use_container_width=True):
        if not new_name.strip():
            st.error("신병 이름을 입력해주세요.")
        else:
            with st.spinner("신병 반영 후 CP-SAT 재계산 중... ⏳"):
                out = rebuild_with_newcomer(new_name.strip(), squad_label)
            if out:
                st.success("✅ 신병이 반영된 새 공정표 생성 완료!")
                if os.path.exists(out):
                    with open(out, "rb") as f:
                        st.download_button(
                            "신병 반영 공정표 다운로드",
                            data=f,
                            file_name=os.path.basename(out),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )


# ---------------------------------------------------
# 6) 근무표 수동 수정
# ---------------------------------------------------
elif page == "근무표 수동 수정":
    st.subheader("근무표 수동 수정 UI (월간 시트)")

    df_month = load_monthly_df()
    if df_month is None:
        st.info("현재 공정표가 없습니다. [근무표 생성] 탭에서 먼저 생성하세요.")
    else:
        st.write(
            "엑셀에서 수정하던 걸 대신해서, 여기서 바로 **주/야/예비/휴가 텍스트를 수정**한 뒤, "
            "엑셀 파일로 다시 저장할 수 있습니다. (요약/공정성 수식은 그대로 유지됩니다.)"
        )

        day_cols = detect_day_columns(df_month)
        edit_cols = ["이름"] + list(day_cols)
        edit_df = df_month[edit_cols].copy()

        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            height=500,
            num_rows="fixed",
            key="manual_edit",
        )

        if st.button("💾 수정 내용 엑셀에 반영해서 새 파일로 저장", use_container_width=True):
            new_path = apply_manual_edits_to_excel(edited)
            if new_path:
                st.success("✅ 수정본 엑셀 저장 완료!")
                with open(new_path, "rb") as f:
                    st.download_button(
                        "수정본 엑셀 다운로드",
                        data=f,
                        file_name=os.path.basename(new_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )


# ---------------------------------------------------
# 7) 디자인 Mock-up
# ---------------------------------------------------
elif page == "디자인 Mock-up":
    st.subheader("Sentinel AIP-lite UI 디자인 Mock-up")

    st.markdown(
        """
        캐치테이블 + Palantir AIP 느낌을 합친 구조:

        - 상단 헤더: 오늘 날짜, 대상 월, 주요 KPI 카드 (총 인원, 생성 여부, 규칙 위반 0건 등)
        - 왼쪽 사이드바: 페이지 네비게이션 + 버전/상태 표시
        - 메인 뷰:
          - 대시보드: 인원별 근무 분포 요약, 야/주 편중도, 주말 예비 분포
          - 근무표 생성: 규칙 설명, 현재 설정, 실행 로그/결과
          - 개별 통계: 선택 인원 타임라인 그래프, 총 근무/예비/휴가 히트맵
          - 신병 투입: Before/After 비교 카드
          - 수동 수정: 표 편집 + 규칙 위반 알림(차후)
        """
    )

    col1, col2 = st.columns([1, 1.2])

    with col1:
        st.markdown("#### Hero 영역")
        st.markdown(
            """
            - 좌측: 서비스 이름 / 설명
            - 우측: 이번 달 목표(예: '야→주 위반 0건', '주말 예비 분산지수 ≤ 2')
            - 상단에 Ops 태그 / 환경 태그(Prod / Test)
            """
        )

    with col2:
        st.markdown("#### Card & Chart 레이아웃")
        st.markdown(
            """
            - 1행: 4개 카드 (총 인원, 공정성 평균, 최대 편차, 위반 규칙 수)
            - 2행: 좌측 큰 그래프(근무 분포), 우측 상세 테이블/필터
            - 색상:
              - 배경: 짙은 네이비 (#020617 ~ #0b1724)
              - 포인트: 민트/하늘색 그라디언트 (#0ea5e9, #22c55e)
              - 서브: CatchTable 느낌의 화이트 카드 가능 (v0.2에서 추가)
            """
        )

    st.info("지금 v0.1은 기능 위주 프로토타입이고, 나중에 색 · 폰트 · 여백을 더 다듬어서 진짜 제품 느낌으로 갈 수 있어.")


